import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook

def converter_xls_para_xlsx(arquivo_entrada, arquivo_saida):
    try:
        livro_xls = xlrd.open_workbook(arquivo_entrada)
        livro_xlsx = Workbook()

        for i in range(livro_xls.nsheets):
            planilha_xls = livro_xls.sheet_by_index(i)
            if i == 0:
                planilha_xlsx = livro_xlsx.active
                planilha_xlsx.title = planilha_xls.name
            else:
                planilha_xlsx = livro_xlsx.create_sheet(title=planilha_xls.name)

            for row in range(planilha_xls.nrows):
                for col in range(planilha_xls.ncols):
                    planilha_xlsx.cell(row=row + 1, column=col + 1, value=planilha_xls.cell_value(row, col))

        livro_xlsx.save(arquivo_saida)
        print(f"Arquivo convertido para: {arquivo_saida}")
        return arquivo_saida

    except Exception as e:
        print(f"Erro ao converter arquivo: {e}")
        return None

def desbloquear_excel(arquivo_entrada, arquivo_saida, senha):
    try:
        wb = load_workbook(arquivo_entrada)

        for sheet in wb.worksheets:
            if sheet.protection.sheet:
                sheet.protection.password = senha
                sheet.protection.sheet = False

        if wb.security.workbookPassword:
            wb.security.workbookPassword = None
            wb.security.lockStructure = False

        wb.save(arquivo_saida)
        print(f"Arquivo desbloqueado e salvo como: {arquivo_saida}")

    except Exception as e:
        print(f"Erro ao desbloquear o arquivo: {e}")

arquivo_entrada = 'estoque distribuidora 08 jan 2024.XLS'
arquivo_convertido = 'estoque_convertido.xlsx'
arquivo_saida = 'estoque_desbloqueado.xlsx'
senha = 'sua_senha_aqui' 

arquivo_convertido = converter_xls_para_xlsx(arquivo_entrada, arquivo_convertido)
if arquivo_convertido:
    desbloquear_excel(arquivo_convertido, arquivo_saida, senha)
